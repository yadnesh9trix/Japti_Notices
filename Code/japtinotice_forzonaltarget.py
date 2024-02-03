import os
import time
# import module
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from datetime import datetime,timedelta
import warnings
warnings.filterwarnings('ignore')
import os
from math import radians, sin, cos, sqrt, atan2
import read_data as rd
import calaculate_distance as cd
from geopy.geocoders import Nominatim
from openpyxl.styles import PatternFill
#-----------------------------------------------------------------------------------------------------------------------
## Define the today's date
# Define today's date
today = datetime.today().date()
# Date format's
tday_dbyfmt = today.strftime("%d_%b_%Y")
tday_dmyfmt = today.strftime("%d%m%Y")
# ---------------------------------------------------------------------------------------------------

## Start
if __name__ == '__main__':
    main_path = r"D:/"
    std_path = r"D:\Master Data/"
    inppath = std_path + "Input/"
    outpth = std_path + "Output/" + tday_dbyfmt + "/"
    paidamount_file = main_path + "Paidamount/"
    tax_data = main_path + "/Tax_Data/"
    os.makedirs(outpth,exist_ok=True)
    mappath = std_path + "Mapping/"

#-----------------------------------------------------------------------------------------
property_data, bill_distributed_details, japtinotice_data,zonemap =  rd.execute_data(inppath ,tax_data)
japti_propertydata =  japtinotice_data.merge(property_data,on='propertycode',how='left')
japti_billsdata = japti_propertydata.merge(bill_distributed_details,on='propertycode',how='left')

data = rd.data_clean(japti_billsdata)

identical_col_df = pd.DataFrame(data, columns=['propertycode', 'Zone', 'Gat', 'own_mobile','Arrears', 'Current Bill', 'Total_Amount',
                                                          'propertyLat','propertyLong','finalusetype',
                                                                      'propertyname', 'propertyaddress','status'])

#--------------------------------------------------------------------------------------------------

usetype_filter = identical_col_df[identical_col_df['finalusetype'].isin(['बिगरनिवासी', 'औद्योगिक'])]
sttaus_np = usetype_filter[usetype_filter['status'].isin(['N' ,'P'])]

# Get current location's latitude and longitude
current_lat, current_lon = cd.get_current_location()

# Calculate distance for each row and add as a new column 'Distance'
sttaus_np['Distance'] = sttaus_np.apply(lambda row: cd.calculate_distance(current_lat, current_lon, row['propertyLat'], row['propertyLong']), axis=1)

# Sort the dataframe based on the 'Distance' column
df_sorted = sttaus_np.sort_values(by='Distance')

# betwn_2L_4L = sttaus_np[(sttaus_np['Total_Amount'] >= 200000) & (sttaus_np['Total_Amount'] < 400000)]
df_above2L = df_sorted[(df_sorted['Total_Amount'] >= 200000)]

#--------------------------------------------------------------------------------------------------
df_above2L['Zone_eng'] = df_above2L['Zone'].map(zonemap)

# lst = [ 'Akurdi', 'Bhosari',  'Nigdi Pradhikaran', 'Talvade', 'Chinchwad', 'Chikhali','Charholi',
#        'Pimpri Nagar', 'Thergaon', 'Fugewadi Dapodi',  'Moshi']
lst = ['Akurdi', 'Bhosari', 'Dighi Bopkhel', 'MNP Bhavan', 'Nigdi Pradhikaran', 'Talvade', 'Chinchwad', 'Chikhali',
       'Pimpri Nagar', 'Thergaon', 'Wakad', 'Kivle', 'Fugewadi Dapodi', 'Pimpri Waghere', 'Sangvi', 'Charholi', 'Moshi']
selctedzone_data = df_above2L[df_above2L['Zone_eng'].isin(lst)]
wout_latlong =  selctedzone_data[selctedzone_data['propertyLong'] > 0]

lst = ['Akurdi', 'Dighi Bopkhel', 'MNP Bhavan',
        'Wakad', 'Kivle',  'Pimpri Waghere', 'Sangvi']
#----------------------------------------------------------------------------------------------------
# lst = ['Akurdi', 'Bhosari', 'Dighi Bopkhel', 'MNP Bhavan', 'Nigdi Pradhikaran', 'Talvade', 'Chinchwad', 'Chikhali',
#        'Pimpri Nagar', 'Thergaon', 'Wakad', 'Kivle', 'Fugewadi Dapodi', 'Pimpri Waghere', 'Sangvi', 'Charholi', 'Moshi']

#-----------------------------------------------------------------------------------------------------
# gatt = wout_latlong['Gat'].unique().tolist()
gatt =  [1, 2, 3, 4, 5, 6, 7, 8, 9, 10,11,12, 13,14, 15, 16,17,18]

selcted_data = pd.DataFrame(wout_latlong, columns=['Zone', 'Gat', 'propertycode',
                                                        'own_mobile', 'Arrears', 'Current Bill', 'Total_Amount',
                                                        'propertyLat','propertyLong','finalusetype',
                                                        'propertyname', 'propertyaddress',
                                                        'Zone_eng'])
rename_data = selcted_data.rename(columns={'Zone':'झोन', 'Gat':'गट क्र', 'propertycode':'मालमत्ता क्रमांक',
                                                        'own_mobile':'मोबाईल क्र.', 'Arrears':'थकबाकी', 'Current Bill':'चालू मागणी रु.', 'Total_Amount':'एकुण मागणी रु.',
                                                        'finalusetype':'वापर प्रकार',
                                                        'propertyname':'मालकाचे नाव', 'propertyaddress':'मालमत्तेचा पत्ता'})

# Function to create Google Maps link
def create_google_maps_link(lat, lon):
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lon}"

# Create a new column 'GoogleMapsLink'
rename_data['GoogleMapsLink'] = rename_data.apply(lambda row: create_google_maps_link(row['propertyLat'], row['propertyLong']), axis=1)

#
# rename_data = pd.DataFrame(rename_data, columns=['झोन','गट क्र','मालमत्ता क्रमांक','वापर प्रकार','मालकाचे नाव','मोबाईल क्र.',
#                                                    'थकबाकी','चालू मागणी रु.','एकुण मागणी रु.','मालमत्तेचा पत्ता',
#                                                    'propertyLat','propertyLong','GoogleMapsLink','Call_Date1','Call_Date2',
#                                                    'Japti_Date','Japti_Status','कारणे',
#                                                     'Zone_eng'])

arrange_data = pd.DataFrame(rename_data, columns=['झोन','गट क्र','मालमत्ता क्रमांक','वापर प्रकार','मालकाचे नाव','मोबाईल क्र.',
                                                   'थकबाकी','चालू मागणी रु.','एकुण मागणी रु.','मालमत्तेचा पत्ता',
                                                   'GoogleMapsLink','Call_Date1','Call_Date2',
                                                   'Japti_Date','Japti_Status','कारणे',
                                                    'Zone_eng'])

arrange_data.to_excel(outpth + "TotalJaptiDataList.xlsx",index=False)
#--------------------------------------------------------------------------------------------------

for i in lst:
    writer = pd.ExcelWriter(outpth + "/" + f"{i}-Japti List.xlsx", engine="xlsxwriter")
    for j in gatt:
        filterdata = arrange_data[(arrange_data['Zone_eng'] == i) & (arrange_data['गट क्र'] == j)]
        if len(filterdata) == 0:
            pass
        else:
            filterdata = filterdata.drop(columns=['Zone_eng'])
            filterdata.to_excel(writer, index=False, sheet_name=f"गट क्र._({str(j)})")

            wb_length = len(filterdata)
            worksheet = writer.sheets[f"गट क्र._({str(j)})"]

            # rule = '"कोर्ट केस,कोर्ट केसस्टे,केंद्रीय सरकारमालमत्ता,राज्य सरकार मालमत्ता,महानगरपालिकेची मालमत्ता,रस्ता रुंदीकरण्यात पडलेली मालमत्ता,''दुबार मालमत्ता,मोकळी जमीन रद्द करणे,बंद कंपनी,पडीक/जीर्ण मालमत्ता,सापडत नसलेली मालमत्ता,BIFR/Liquidation,इतर,"'
            # rule = '"Yes,No"'
            # dropdown_range = f'R2:R{wb_length + 1}'
            # worksheet.data_validation(dropdown_range, {'validate': 'list', 'source': rule})
            worksheet.freeze_panes(1, 3)
            workbook = writer.book

            worksheet.set_column('C1:O1', 13)
            # worksheet.set_column('D1:D1', 16)
            border_format = workbook.add_format({'border': 1,
                                                 'align': 'left',
                                                 'font_color': '#000000',
                                                 'font_size': 20})
            worksheet.conditional_format(f'A1:T{wb_length + 1}', {'type': 'cell',
                                                                  'criteria': '>=',
                                                                  'text_wrap': True,
                                                                  'value': 0,
                                                                  'format': border_format})
            worksheet.set_row(wb_length + 1, 28)
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

            worksheet.conditional_format(f'I2:I{wb_length + 1}', {'type': 'cell',
                                                                  'criteria': '>=',
                                                                  'value': 400000,
                                                                  'format': workbook.add_format({'bg_color': 'red', 'font_color': 'white'})})
            worksheet.conditional_format(f'I2:I{wb_length + 1}', {'type': 'cell',
                                                                  'criteria': '<',
                                                                  'value': 400000,
                                                                   'format': workbook.add_format({'bg_color': 'blue', 'font_color': 'white'})})
    writer.close()


